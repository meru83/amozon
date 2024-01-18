import mysql.connector
import sys
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
import csv


# ########################################
# コマンドライン引数から「seller_id」取得
# ########################################
args = sys.argv

# seller_id = args[1]
seller_id = args[1] if len(args) > 1 else None

# ########################################
# データベースから対象データ取得
# ########################################

# DBへ接続
conn = mysql.connector.connect(
    user='root',
    password='',
    host='localhost',
    database='complete'
)

# DBの接続確認
if not conn.is_connected():
    raise Exception("MySQLサーバへの接続に失敗しました")

# 取得結果を辞書型で扱う設定
cur = conn.cursor(dictionary=True)

# SQL文字列
query__for_fetching = """
SELECT
    p.seller_id,
    DATE_FORMAT(o.create_at, '%Y') AS year,
    DATE_FORMAT(o.create_at, '%Y-%m') AS month,
    SUM(d.detail_total) AS total
FROM
    orders_detail d
LEFT JOIN
    orders o ON (d.order_id = o.order_id)
LEFT JOIN
    products p ON (d.product_id = p.product_id)
WHERE
    o.order_status = "配達完了"
AND
    p.seller_id = "{seller_id}"
GROUP BY
    p.seller_id, month
ORDER BY
    p.seller_id, month
;
""".format(seller_id=seller_id)

# SQL実行
cur.execute(query__for_fetching)

# SQL実行結果をCSVファイルに書き込み
with open('sample.csv', 'w', encoding="utf8") as f :
    f.write(f'年,月,売上\n')
    for fetched_line in cur.fetchall():
        seller_id = fetched_line['seller_id']
        year = fetched_line['month'][0:4]
        month = fetched_line['month'][5:7]
        total = fetched_line['total']
        f.write(f'{year},{month},{total}\n')


def getYears(df) :
    returnYears = []
    tempYear = 0

    for index, row in df.iterrows():
        if(tempYear != row['年']) :
            tempYear = row['年']
            returnYears.append(row['年'])

    return returnYears

# #######################################
# Excelファイル作成
# #######################################
# CSVファイルの読み込み
csv_file = 'sample.csv'  # CSVファイルのパス
# data = pd.read_csv(csv_file)

# 空のリストを用意
data = []

# CSVファイルを開き、内容を読み込む
with open(csv_file, newline='', encoding='utf-8') as file:
    reader = csv.reader(file)
    for row in reader:
        data.append(row)

# DataFrameに変換
df = pd.DataFrame(data[1:], columns=data[0])

# ワークブックの作成
wb = Workbook()
wb.remove(wb.active)  # デフォルトのシートを削除

# 年度ごとにデータを分割し、シートを作成
for year in df['年'].unique():
    ws = wb.create_sheet(title=str(year))
    year_data = df[df['年'] == year]

    # 「売上」列が文字列になっているので、数値に変換
    year_data['売上'] = pd.to_numeric(year_data['売上'])

    # データをExcelシートに書き込む
    for r in dataframe_to_rows(year_data, index=False, header=True):
        print(r)
        ws.append(r)

    rmin = ws.min_row
    rmax = ws.max_row
    cmin = ws.min_column
    cmax = ws.max_column

    chart = LineChart()
    # X : 月
    src = Reference(ws, min_col=cmin+2, min_row=rmin+1, max_col=cmax, max_row=rmax)
    chart.add_data(src, titles_from_data=False)

    # Y : 売上
    cat = Reference(ws, min_col=cmin+1, min_row=rmin+1, max_row=rmax)  # 項目名の設定
    chart.set_categories(cat)

    chart.title = seller_id  # グラフタイトル
    chart.x_axis.title = '月別'  # 軸ラベル
    chart.y_axis.title = '売上'
    chart.anchor = 'E8'  # グラフの表示位置
    chart.legend = None
    chart.width = 16  # グラフのサイズ
    chart.height = 8
    ws.add_chart(chart)
try:
    # 既存のコード
    # ★★★ 変更ここから ★★★
    # wb.save('C:/xampp/htdocs/php/amo/py/excel/rireki.xlsx')
    wb.save('rireki.xlsx')
    # ★★★ 変更ここまで ★★★
except Exception as e:
    print(f"エラー: {e}")
#python rireki.py rion_bank